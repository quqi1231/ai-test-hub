"""
接口导入服务 - 支持 Excel, CSV, YAML, JSON 格式
"""
import json
import yaml
import csv
import io
from typing import List, Dict, Any
from openpyxl import load_workbook


def parse_json(content: str) -> List[Dict[str, Any]]:
    """解析 JSON 格式"""
    data = json.loads(content)
    # 支持单接口和接口列表
    if isinstance(data, dict):
        if "interfaces" in data:
            return data["interfaces"]
        return [data]
    elif isinstance(data, list):
        return data
    return []


def parse_yaml(content: str) -> List[Dict[str, Any]]:
    """解析 YAML 格式"""
    data = yaml.safe_load(content)
    if isinstance(data, dict):
        if "interfaces" in data:
            return data["interfaces"]
        return [data]
    elif isinstance(data, list):
        return data
    return []


def parse_csv(content: str) -> List[Dict[str, Any]]:
    """解析 CSV 格式"""
    interfaces = []
    reader = csv.DictReader(io.StringIO(content))
    for row in reader:
        # 清理空值
        interface = {k: v for k, v in row.items() if v}
        # 解析 JSON 字段
        for field in ['headers', 'params', 'body']:
            if field in interface and interface[field]:
                try:
                    interface[field] = json.loads(interface[field])
                except:
                    pass
        interfaces.append(interface)
    return interfaces


def parse_excel(file_content: bytes) -> List[Dict[str, Any]]:
    """解析 Excel 格式"""
    interfaces = []
    wb = load_workbook(file_content)
    ws = wb.active
    
    # 获取表头
    headers = [cell.value for cell in ws[1]]
    
    # 读取数据行
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        interface = {}
        for i, value in enumerate(row):
            if i < len(headers) and value:
                key = headers[i]
                # 解析 JSON 字段
                if key in ['headers', 'params', 'body']:
                    try:
                        value = json.loads(str(value))
                    except:
                        pass
                interface[key] = value
        if interface:
            interfaces.append(interface)
    
    return interfaces


def normalize_interface(data: Dict[str, Any]) -> Dict[str, Any]:
    """标准化接口数据"""
    return {
        "name": data.get("name", data.get("接口名称", "未命名")),
        "method": data.get("method", data.get("方法", "GET")).upper(),
        "url": data.get("url", data.get("path", data.get("接口地址", ""))),
        "description": data.get("description", data.get("描述", "")),
        "headers": data.get("headers", data.get("请求头", {})),
        "params": data.get("params", data.get("参数", {})),
        "body": data.get("body", data.get("请求体", {})),
    }


def parse_postman_collection(content: str) -> List[Dict[str, Any]]:
    """解析 Postman 集合格式 (v2.1)"""
    data = json.loads(content)
    interfaces = []
    
    # Postman v2.1 格式
    if "item" in data:
        items = data["item"]
        
        def parse_items(items: List, folder_name: str = ""):
            for item in items:
                # 如果是文件夹（包含 item）
                if "item" in item:
                    folder = item.get("name", folder_name)
                    parse_items(item["item"], folder)
                # 如果是请求
                elif "request" in item:
                    req = item["request"]
                    url_data = req.get("url", {})
                    
                    # 处理 URL（可能是字符串或对象）
                    url = ""
                    if isinstance(url_data, str):
                        url = url_data
                    elif isinstance(url_data, dict):
                        raw = url_data.get("raw", "")
                        protocol = url_data.get("protocol", "https")
                        host = ".".join(url_data.get("host", []))
                        path = "/".join(url_data.get("path", []))
                        url = f"{protocol}://{host}/{path}" if host else raw
                    
                    # 处理请求体
                    body = None
                    body_type = "json"
                    if "body" in req:
                        body_data = req["body"]
                        if body_data:
                            mode = body_data.get("mode", "raw")
                            if mode == "raw":
                                body = body_data.get("raw")
                                body_type = body_data.get("options", {}).get("raw", {}).get("language", "json")
                            elif mode == "formdata":
                                body = {k: v.get("value") for k, v in body_data.get("formdata", {}).items()}
                                body_type = "form-data"
                            elif mode == "urlencoded":
                                body = {k: v.get("value") for k, v in body_data.get("urlencoded", {}).items()}
                                body_type = "x-www-form-urlencoded"
                    
                    # 处理 Headers
                    headers = {}
                    for h in req.get("header", []):
                        if h.get("key"):
                            headers[h["key"]] = h.get("value", "")
                    
                    interfaces.append({
                        "name": item.get("name", "未命名"),
                        "method": req.get("method", "GET").upper(),
                        "url": url,
                        "description": req.get("description", ""),
                        "headers": headers,
                        "params": {},
                        "body": body,
                        "body_type": body_type,
                        "folder": folder_name
                    })
        
        parse_items(items)
    
    return interfaces


def parse_swagger(content: str) -> List[Dict[str, Any]]:
    """解析 Swagger/OpenAPI 格式"""
    data = json.loads(content) if isinstance(content, str) else content
    interfaces = []
    
    # OpenAPI 3.x
    paths = data.get("paths", {})
    servers = data.get("servers", [])
    base_url = servers[0].get("url", "") if servers else ""
    
    for path, methods in paths.items():
        for method, details in methods.items():
            if method.upper() not in ["GET", "POST", "PUT", "DELETE", "PATCH", "OPTIONS", "HEAD"]:
                continue
            
            # 解析请求体
            body = None
            body_type = "json"
            request_body = details.get("requestBody", {})
            if request_body:
                content = request_body.get("content", {})
                if "application/json" in content:
                    schema = content["application/json"].get("schema", {})
                    body = {"type": schema.get("type", "object")}
                    body_type = "json"
            
            # 解析参数
            params = {}
            for param in details.get("parameters", []):
                if param.get("in") == "query":
                    params[param.get("name", "")] = ""
            
            # 解析响应
            responses = details.get("responses", {})
            
            interfaces.append({
                "name": details.get("summary", details.get("operationId", path)),
                "method": method.upper(),
                "url": base_url + path,
                "description": details.get("description", ""),
                "headers": {},
                "params": params,
                "body": body,
                "body_type": body_type,
                "tags": details.get("tags", [])
            })
    
    return interfaces


def detect_format(content: str) -> str:
    """自动检测导入格式"""
    try:
        data = json.loads(content)
        
        # Postman 格式检测
        if "item" in data:
            return "postman"
        
        # Swagger/OpenAPI 格式检测
        if "openapi" in data or "swagger" in data:
            return "swagger"
        
        # 普通 JSON 格式
        if isinstance(data, (dict, list)):
            return "json"
            
    except json.JSONDecodeError:
        pass
    
    # 尝试 YAML
    try:
        yaml.safe_load(content)
        return "yaml"
    except:
        pass
    
    return "json"


def import_interfaces(file_content: bytes, filename: str, project_id: int) -> List[Dict[str, Any]]:
    """导入接口主函数 - 支持自动格式检测"""
    content = file_content.decode('utf-8')
    ext = filename.lower().split('.')[-1]
    
    # 自动检测格式
    if ext == 'json':
        format_type = detect_format(content)
    else:
        format_type = ext
    
    # 根据格式解析
    if format_type == 'postman':
        raw_interfaces = parse_postman_collection(content)
    elif format_type == 'swagger':
        raw_interfaces = parse_swagger(content)
    elif format_type in ['xlsx', 'xls']:
        raw_interfaces = parse_excel(file_content)
    elif format_type == 'csv':
        raw_interfaces = parse_csv(content)
    elif format_type in ['yaml', 'yml']:
        raw_interfaces = parse_yaml(content)
    elif format_type == 'json':
        raw_interfaces = parse_json(content)
    else:
        raise ValueError(f"不支持的文件格式: {ext}")
    
    # 标准化并添加 project_id
    interfaces = []
    for data in raw_interfaces:
        interface = normalize_interface(data)
        interface["project_id"] = project_id
        interfaces.append(interface)
    
    return interfaces
